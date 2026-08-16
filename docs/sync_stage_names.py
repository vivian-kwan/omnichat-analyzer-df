#!/usr/bin/env python3
"""
Sync the "Stage Names" sheet in omnichat-copy-inventory.xlsx -> every place a
classification stage name is a hardcoded identifier (not just display copy):

  - Plug-in/content.js        LABEL_COLORS key, LABEL_ORDER element,
                              TEMPLATE_STAGES element, the Quotation chip-class
                              check, the <option value="..."> pair in the
                              Templates-tab stage picker
  - routes/templates.js       VALID_STAGES element
  - data/templates.json       migrates any saved template's "stage" field
  - AGENTS/skills/home-analysis-skill.md   the AI prompt's category name,
  - data/skills/home-analysis.txt          wherever it's mentioned

This is a different, riskier operation than sync_copy_from_xlsx.py — it
rewrites the live AI classification prompt. Always --dry-run first.

Usage:
    python3 sync_stage_names.py [--dry-run] [--xlsx PATH]
"""
import argparse
import json
import re
from pathlib import Path

import openpyxl

DEV_ROOT = Path(
    "/Users/user/Library/CloudStorage/GoogleDrive-google@viviankwan.work/"
    "Shared drives/DF/DF 創意家居/VIV WIP/Omnichat/omnichat-analyzer-v2/development"
)
REPO_ROOT = Path(__file__).resolve().parent.parent

CONTENT_JS = DEV_ROOT / "Plug-in" / "content.js"
SKILL_MD = DEV_ROOT / "AGENTS" / "skills" / "home-analysis-skill.md"
TEMPLATES_ROUTE = REPO_ROOT / "routes" / "templates.js"
TEMPLATES_JSON = REPO_ROOT / "data" / "templates.json"
SKILL_TXT = REPO_ROOT / "data" / "skills" / "home-analysis.txt"

def current_stages():
    """Read the live LABEL_ORDER from content.js — the actual current set of
    valid stage names — rather than a hardcoded list, so a stage that was
    already renamed once can be renamed again."""
    content = CONTENT_JS.read_text(encoding="utf-8")
    m = re.search(r'const LABEL_ORDER = \[(.*?)\];', content)
    names = re.findall(r'"([^"]+)"', m.group(1)) if m else []
    return [n for n in names if n != "Unlabeled"]


def load_renames(xlsx_path):
    """Columns: Type | Tag/Stage | Font Role | Font Spec | Value | New Value.
    Only Type=="Stage" rows apply here — Tag/Stage is the current name,
    New Value (col F) is the rename target."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Copy & Stage Names"]
    renames = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_type, old, new = row[0], row[1], row[5]
        if row_type != "Stage" or not old or not new:
            continue
        old, new = str(old).strip(), str(new).strip()
        if old != new:
            renames.append((old, new))
    # Longest OLD name first, so "Quotation start" is handled before the
    # standalone "Quotation" rename could accidentally eat into it.
    renames.sort(key=lambda p: -len(p[0]))
    return renames


def replace_on_matching_line(content, line_contains, old, new, quote='"'):
    """Replace old->new (quoted) only on the one line containing line_contains."""
    lines = content.split("\n")
    changed = 0
    for i, line in enumerate(lines):
        if line_contains in line and f'{quote}{old}{quote}' in line:
            lines[i] = line.replace(f'{quote}{old}{quote}', f'{quote}{new}{quote}')
            changed += 1
    return "\n".join(lines), changed


def sync_content_js(old, new, dry_run):
    content = CONTENT_JS.read_text(encoding="utf-8")
    original = content
    report = []

    content, n = replace_on_matching_line(content, "const LABEL_COLORS", old, new)
    report.append(f"LABEL_COLORS key: {n} change(s)")

    content, n = replace_on_matching_line(content, "const LABEL_ORDER", old, new)
    report.append(f"LABEL_ORDER element: {n} change(s)")

    content, n = replace_on_matching_line(content, "const TEMPLATE_STAGES", old, new)
    report.append(f"TEMPLATE_STAGES element: {n} change(s)")

    if old == "Quotation":
        content, n = replace_on_matching_line(content, "g.label === ", old, new)
        report.append(f"Quotation chip-class check: {n} change(s)")

    pattern = re.compile(r'value="%s">%s<' % (re.escape(old), re.escape(old)))
    content, n = pattern.subn('value="%s">%s<' % (new, new), content)
    report.append(f"<option> stage picker: {n} change(s)")

    if content != original and not dry_run:
        CONTENT_JS.write_text(content, encoding="utf-8")
    return report, content != original


def sync_templates_route(old, new, dry_run):
    content = TEMPLATES_ROUTE.read_text(encoding="utf-8")
    original = content
    content, n = replace_on_matching_line(content, "const VALID_STAGES", old, new, quote="'")
    if content != original and not dry_run:
        TEMPLATES_ROUTE.write_text(content, encoding="utf-8")
    return [f"VALID_STAGES element: {n} change(s)"], content != original


def sync_templates_json(old, new, dry_run):
    templates = json.loads(TEMPLATES_JSON.read_text(encoding="utf-8"))
    migrated = 0
    for t in templates:
        if t.get("stage") == old:
            t["stage"] = new
            migrated += 1
    if migrated and not dry_run:
        TEMPLATES_JSON.write_text(json.dumps(templates, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    return [f"saved templates migrated: {migrated}"], migrated > 0


def sync_prose_file(path, old, new, dry_run):
    content = path.read_text(encoding="utf-8")
    original = content
    # Word-boundary match; guard against "Quotation" eating into "Quotation start".
    guard = r"(?!\s+start\b)" if old == "Quotation" else ""
    pattern = re.compile(r'\b%s\b%s' % (re.escape(old), guard))
    content, n = pattern.subn(new, content)
    if content != original and not dry_run:
        path.write_text(content, encoding="utf-8")
    return [f"{path.name}: {n} occurrence(s) replaced"], content != original


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--xlsx", default=str(Path(__file__).parent / "omnichat-copy-inventory.xlsx"))
    args = ap.parse_args()

    renames = load_renames(args.xlsx)
    if not renames:
        print("No renames pending — every row's New Name matches Current Name.")
        return

    valid_stages = current_stages()
    for old, new in renames:
        if old not in valid_stages:
            print(f"⚠ Skipping {old!r} — not found in content.js's current LABEL_ORDER ({valid_stages}). "
                  f"Did the 'Current Name' column get out of sync with the actual code?")
            continue
        print(f"\n=== Renaming {old!r} -> {new!r} ===")
        for label, fn, args_ in [
            ("content.js", sync_content_js, (old, new, args.dry_run)),
            ("routes/templates.js", sync_templates_route, (old, new, args.dry_run)),
            ("data/templates.json", sync_templates_json, (old, new, args.dry_run)),
            ("AGENTS/skills/home-analysis-skill.md", sync_prose_file, (SKILL_MD, old, new, args.dry_run)),
            ("data/skills/home-analysis.txt", sync_prose_file, (SKILL_TXT, old, new, args.dry_run)),
        ]:
            if fn is sync_prose_file:
                report, changed = fn(*args_)
            else:
                report, changed = fn(*args_)
            print(f"  [{label}]")
            for line in report:
                print(f"    - {line}")

    print(f"\n{'DRY RUN — nothing written' if args.dry_run else 'Applied.'}")
    if not args.dry_run:
        # Write Tag/Stage = New Value back into the sheet so re-running is a no-op.
        wb = openpyxl.load_workbook(args.xlsx)
        ws = wb["Copy & Stage Names"]
        rename_map = dict(renames)
        for row in ws.iter_rows(min_row=2):
            if row[0].value != "Stage":
                continue
            old_val = row[1].value
            if old_val in rename_map:
                row[1].value = rename_map[old_val]
        wb.save(args.xlsx)
        print("'Copy & Stage Names' sheet's Tag/Stage column updated to match.")


if __name__ == "__main__":
    main()
